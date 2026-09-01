from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from core import db, new_id, now_iso, gen_maintenance_no

SEED_FILE = Path(__file__).parent / "seed_data" / "dashboard.xlsx"


def to_iso_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date().isoformat()
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_row(rows):
    """Find the row that looks like a header (contains 'SAP')."""
    for idx, r in enumerate(rows[:5]):
        joined = " ".join(_s(c) for c in r).lower()
        if "sap" in joined:
            return idx
    return 1


def parse_workbook(path_or_bytes):
    wb = openpyxl.load_workbook(path_or_bytes, data_only=True, read_only=True)
    equipment, maintenance = [], []
    names = {n.lower(): n for n in wb.sheetnames}

    eqp_sheet = names.get("eqp register")
    if eqp_sheet:
        rows = list(wb[eqp_sheet].iter_rows(values_only=True))
        h = _header_row(rows)
        for r in rows[h + 1:]:
            sap = _s(r[1] if len(r) > 1 else "")
            name = _s(r[3] if len(r) > 3 else "")
            if not sap and not name:
                continue
            equipment.append({
                "sap_no": sap,
                "mfg_no": _s(r[2] if len(r) > 2 else ""),
                "name": name,
                "category": _s(r[4] if len(r) > 4 else ""),
                "manufacturer": _s(r[5] if len(r) > 5 else ""),
                "date_of_purchase": to_iso_date(r[6] if len(r) > 6 else None),
                "physical_condition": _s(r[7] if len(r) > 7 else ""),
            })

    src_sheet = names.get("source data")
    if src_sheet:
        rows = list(wb[src_sheet].iter_rows(values_only=True))
        h = _header_row(rows)
        for r in rows[h + 1:]:
            r = list(r) + [None] * (16 - len(r))
            sap = _s(r[1])
            problem = _s(r[4])
            if not sap and not problem:
                continue
            maintenance.append({
                "maintenance_date": to_iso_date(r[0]),
                "sap_no": sap,
                "equipment_name": _s(r[2]),
                "category": _s(r[3]),
                "problem_damage": problem,
                "type_of_maintenance": _s(r[5]),
                "action_taken": _s(r[6]),
                "date_closed": to_iso_date(r[7]),
                "duration_days": _s(r[8]),
                "lead_technician": _s(r[9]),
                "checked_by": _s(r[10]),
                "final_condition": _s(r[11]),
                "pending_maintenance": _s(r[12]),
                "progress_update": _s(r[13]),
                "maintenance_category": _s(r[15]),
            })
    wb.close()
    return equipment, maintenance


async def _insert_equipment(row, source="import"):
    doc = {
        "id": new_id(),
        "sap_no": row["sap_no"],
        "mfg_no": row.get("mfg_no", ""),
        "name": row.get("name", ""),
        "category": row.get("category", ""),
        "manufacturer": row.get("manufacturer", ""),
        "date_of_purchase": row.get("date_of_purchase"),
        "physical_condition": row.get("physical_condition", ""),
        "placement": "Base",
        "placement_detail": "Base",
        "operational_status": "Operational",
        "current_job_id": None,
        "current_client_id": None,
        "source": source,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.equipment.insert_one(doc)
    await db.location_history.insert_one({
        "id": new_id(), "equipment_id": doc["id"], "from_placement": None,
        "to_placement": "Base", "placement_detail": "Base", "job_id": None,
        "reason": "Initial registration", "created_by": "system", "created_at": now_iso(),
    })
    return doc


async def _insert_maintenance(row, equipment_map):
    sap = row["sap_no"]
    eq = equipment_map.get(sap)
    if not eq:
        eq = await _insert_equipment({
            "sap_no": sap, "mfg_no": "", "name": row.get("equipment_name", ""),
            "category": row.get("category", ""), "manufacturer": "",
            "date_of_purchase": None, "physical_condition": "",
        }, source="import")
        equipment_map[sap] = eq
    try:
        dur = int(float(row.get("duration_days") or 0))
    except Exception:
        dur = 0
    mnt_no = await gen_maintenance_no()
    closed = bool(row.get("date_closed"))
    doc = {
        "id": new_id(),
        "mnt_no": mnt_no,
        "equipment_id": eq["id"],
        "sap_no": sap,
        "equipment_name": eq.get("name"),
        "maintenance_date": row.get("maintenance_date"),
        "date_closed": row.get("date_closed"),
        "maintenance_category": row.get("maintenance_category", ""),
        "type_of_maintenance": row.get("type_of_maintenance", ""),
        "problem_damage": row.get("problem_damage", ""),
        "failure_found": row.get("problem_damage", ""),
        "root_cause": "",
        "action_taken": row.get("action_taken", ""),
        "current_status": row.get("problem_damage", ""),
        "duration_days": dur,
        "lead_technician": row.get("lead_technician", ""),
        "support_technicians": [],
        "checked_by": row.get("checked_by", ""),
        "final_condition": row.get("final_condition", ""),
        "remark": row.get("final_condition", ""),
        "pending_maintenance": row.get("pending_maintenance", ""),
        "progress_update": row.get("progress_update", ""),
        "client_id": None,
        "client_name": None,
        "job_id": None,
        "job_number": None,
        "parts_consumed": [],
        "attachments": [],
        "status": "Closed" if closed else "Open",
        "source": "import",
        "created_by": "system",
        "created_at": now_iso(),
        "closed_at": row.get("date_closed"),
    }
    await db.maintenance.insert_one(doc)
    # Failure record derived from the problem/current status
    fname = (row.get("problem_damage") or "").strip()
    if fname:
        await db.failures.insert_one({
            "id": new_id(), "equipment_id": eq["id"], "maintenance_id": doc["id"],
            "mnt_no": mnt_no, "failure_name": fname, "description": row.get("action_taken", ""),
            "root_cause": "", "occurred_date": row.get("maintenance_date"),
            "created_at": now_iso(),
        })
    if not closed:
        await db.equipment.update_one({"id": eq["id"]},
                                      {"$set": {"operational_status": "Under Maintenance"}})
    return doc


DEFAULT_INVENTORY = [
    ("SP-0001", "Bearing Set (Decanter)", "Spare Part", "BRG-DEC-14", "SET", 12, 4, "Workshop Rack A1"),
    ("SP-0002", "Mechanical Seal Kit", "Spare Part", "SEAL-KIT-01", "SET", 8, 3, "Workshop Rack A2"),
    ("SP-0003", "Pressure Transmitter", "Spare Part", "PT-4520", "EA", 5, 2, "Workshop Rack B1"),
    ("SP-0004", "UPS Battery 12V", "Spare Part", "UPS-BAT-12", "EA", 6, 2, "Workshop Rack B2"),
    ("SP-0005", "Sun Gear Assembly", "Spare Part", "SG-1000", "EA", 3, 1, "Workshop Rack C1"),
    ("SP-0006", "O-Ring Kit", "Spare Part", "ORK-STD", "SET", 20, 6, "Workshop Rack C2"),
    ("CN-0001", "Contact Cleaner", "Consumable", "CLN-CAN-400", "CAN", 24, 8, "Store Shelf 1"),
    ("CN-0002", "Grease EP2", "Consumable", "GRS-EP2", "KG", 15, 5, "Store Shelf 2"),
    ("CN-0003", "Gear Oil ISO220", "Consumable", "OIL-220", "LTR", 40, 10, "Store Shelf 3"),
    ("CN-0004", "Loctite 243", "Consumable", "LCT-243", "EA", 10, 3, "Store Shelf 4"),
]


async def seed_inventory():
    if await db.inventory_items.count_documents({}) > 0:
        return
    for code, name, typ, pn, unit, stock, mn, loc in DEFAULT_INVENTORY:
        await db.inventory_items.insert_one({
            "id": new_id(), "item_code": code, "item_name": name, "type": typ,
            "part_number": pn, "unit": unit, "stock": stock, "min_stock": mn,
            "storage_location": loc, "created_at": now_iso(), "updated_at": now_iso(),
        })


async def seed_demo_clients():
    if await db.clients.count_documents({}) > 0:
        return
    c = {"id": new_id(), "name": "ABC Energy", "code": "ABC", "contact": "ops@abcenergy.com",
         "notes": "Offshore drilling client", "created_at": now_iso()}
    await db.clients.insert_one(c)


async def seed_from_excel():
    if await db.equipment.count_documents({}) > 0:
        return {"skipped": True}
    equipment_rows, maintenance_rows = parse_workbook(str(SEED_FILE))
    equipment_map = {}
    for row in equipment_rows:
        if row["sap_no"] and row["sap_no"] in equipment_map:
            continue
        doc = await _insert_equipment(row, source="excel-seed")
        if row["sap_no"]:
            equipment_map[row["sap_no"]] = doc
    # maintenance chronological (oldest first so MNT numbers ascend by date)
    maintenance_rows.sort(key=lambda r: r.get("maintenance_date") or "")
    for row in maintenance_rows:
        await _insert_maintenance(row, equipment_map)
    await seed_inventory()
    await seed_demo_clients()
    return {"equipment": len(equipment_map), "maintenance": len(maintenance_rows)}
